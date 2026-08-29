"""
Medición del evaluador contra un conjunto de referencia etiquetado a mano.

Por qué existe
--------------
Hasta ahora lo único que se podía afirmar del lector era cuántos documentos
producían veredicto: «lee 7 de 13». Eso no dice si lo que lee es correcto. Un
lector que se invente siete fechas plausibles saca la misma cifra que uno que las
lea bien, y en un evaluador esa diferencia lo es todo — porque una fecha
inventada no se queda en un hueco: acusa a un compañero de un fallo que no ha
cometido.

El conjunto de referencia lo etiqueta una persona leyendo los documentos. Contra
él se miden **cinco desenlaces por campo**, y los cinco son distintos a propósito:

  · acierto             — el sistema dice lo mismo que la persona
  · omisión             — la persona lo leyó y el sistema no supo
  · ERROR               — los dos dicen algo y no es lo mismo
  · abstención correcta — el documento no lo dice y el sistema calla
  · INVENCIÓN           — el documento no lo dice y el sistema lo afirma

Los dos en mayúsculas son los que le quitan autoridad al evaluador. Una omisión
es una limitación declarada y se ve; un error y una invención se disfrazan de
dato bueno. Por eso se cuentan aparte y por eso la cifra que se lleva a la
memoria no es «cuántos lee» sino **cuántos lee bien y cuántas veces se calla
cuando debe**.

Uso
---
    python medir.py [referencia/etiquetado.xlsx] [--modo determinista|asistido]
"""

import re
import sys
from datetime import date, datetime
from pathlib import Path

from nucleo import pdf as P
from modulos import vigencia as V

RAIZ_DOCS = Path("demo/datos/vigencia")
FECHA = date(2026, 8, 28)

# Los campos que se miden, con el nombre que tienen en la hoja y en el sistema.
CAMPOS = [
    ("fecha_firma", "fecha_emision", "fecha"),
    ("fecha_inicio", "fecha_inicio", "fecha"),
    ("plazo_anios", "anios_pactados", "numero"),
    ("fecha_vencimiento", "fecha_caducidad", "fecha"),
    ("prorroga", "prorroga_tipo", "texto"),
    ("preaviso_dias", "_preaviso", "numero"),
]

# Cómo se escribe cada campo en la columna «no consta»: la persona escribe en
# castellano corriente —«preaviso dias», «plazo años»— y no con el nombre técnico.
ALIAS_NO_CONSTA = {
    "fecha_firma": ["firma", "fecha de firma"],
    "fecha_inicio": ["inicio", "fecha de inicio"],
    "plazo_anios": ["plazo", "plazo anos", "plazo anios", "anos", "duracion"],
    "fecha_vencimiento": ["vencimiento", "fecha de vencimiento", "caducidad"],
    "prorroga": ["prorroga"],
    "preaviso_dias": ["preaviso", "preaviso dias"],
}

# Vocabulario de estados. La persona escribe «titulo consumado» y «no clasificado
# (revisar)»; el sistema emite identificadores.
#
# Aquí vivió un rato `depende_de_otro`, que el sistema emitía para los anexos y
# se aceptaba como equivalente de «no clasificado». Se retiró el 28/08 de los dos
# sitios: el vocabulario de esta rama es el de la prueba inicial de Martín, y una
# medición que acepta un estado inventado por el evaluador mide contra una regla
# que nadie acordó.
ESTADOS_EQUIV = {
    "vigente": {"vigente"},
    "caducado": {"caducado", "vencido"},
    "obsoleto": {"obsoleto", "sustituido"},
    "titulo_consumado": {"titulo consumado", "titulo_consumado", "consumado"},
    "no_clasificado": {"no clasificado", "no_clasificado", "revisar",
                       "vigencia no determinada", "no determinado"},
    "no_aplica_vigencia": {"no aplica", "no_aplica_vigencia", "sin vigencia"},
}
# El sistema puede decir esto y contar como acierto de lo que la persona escribió.
MAS_PRECISO_QUE = {
    "no_aplica_vigencia": "no_clasificado",
}


def _plano(s):
    import unicodedata
    s = unicodedata.normalize("NFKD", str(s or "").lower())
    s = "".join(c for c in s if not unicodedata.combining(c))
    return re.sub(r"[^a-z0-9 ,]+", " ", s).strip()


def normalizar_estado(v):
    # El paréntesis se corta ANTES de aplanar: «no clasificado (revisar)» es el
    # estado más la duda de quien etiquetó, y la duda no forma parte del estado.
    # Aplanar primero se comía el paréntesis y dejaba «no clasificado revisar»,
    # que no casaba con nada — y contaba como fallo del sistema una diferencia de
    # puntuación en la hoja de cálculo.
    t = _plano(str(v or "").split("(")[0])
    for canon, formas in ESTADOS_EQUIV.items():
        if t in {_plano(f) for f in formas} or t == canon:
            return canon
    return t or None


def a_fecha(v):
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    from nucleo.texto import fecha_de
    return fecha_de(str(v))


def a_numero(v):
    if v is None or v == "":
        return None
    m = re.search(r"\d+", str(v))
    return int(m.group(0)) if m else None


def leer_referencia(ruta):
    """
    Filas etiquetadas de verdad, y sólo ésas.

    Una fila cuenta como referencia **cuando tiene el estado escrito**. El resto
    conserva los valores que propuso el sistema y medirse contra ellos sería
    darse la razón a uno mismo: el error más silencioso que puede cometer una
    evaluación.
    """
    from openpyxl import load_workbook
    ws = load_workbook(ruta, data_only=True)["Referencia"]
    cab = {ws.cell(row=6, column=c).value: c for c in range(1, 20)
           if ws.cell(row=6, column=c).value}
    filas = []
    for r in range(8, ws.max_row + 1):
        doc = ws.cell(row=r, column=cab["documento"]).value
        if not doc:
            continue
        estado = ws.cell(row=r, column=cab["estado_correcto"]).value
        if not estado:
            continue                      # sin etiquetar: no se mide
        fila = {"documento": str(doc).strip(),
                "estado": normalizar_estado(estado),
                "familia": _plano(ws.cell(row=r, column=cab["familia_correcta"]).value),
                "notas": ws.cell(row=r, column=cab["notas"]).value or ""}
        for col, _, tipo in CAMPOS:
            v = ws.cell(row=r, column=cab[col]).value
            fila[col] = a_fecha(v) if tipo == "fecha" else (
                a_numero(v) if tipo == "numero" else (_plano(v) or None))
        crudo = _plano(ws.cell(row=r, column=cab["no_consta_en_el_documento"]).value)
        trozos = [t.strip() for t in crudo.split(",") if t.strip()]
        fila["no_consta"] = {
            campo for campo, _n, _t in CAMPOS
            if any(t == _plano(a) or _plano(a) in t
                   for t in trozos for a in ALIAS_NO_CONSTA[campo] + [campo])}
        filas.append(fila)
    return filas


def leer_sistema(nombres, modo="determinista"):
    rutas = []
    for n in nombres:
        p = RAIZ_DOCS / f"{n}.pdf"
        rutas.append(p if p.is_file() else None)
    docs = [P.leer(p) for p in rutas if p]
    esp, ctx = V.verdad_de_campo(docs, FECHA, modo=modo)
    return {e["id_documento"]: e for e in esp}, ctx


def comparar(ref, sis):
    """Los cinco desenlaces, por campo."""
    out = []
    for col, campo, tipo in CAMPOS:
        esperado = ref[col]
        no_consta = col in ref["no_consta"]
        c = sis["campos"] if sis else {}
        if campo == "_preaviso":
            obtenido = (c.get("antelacion") or {}).get("dias")
        else:
            obtenido = c.get(campo)
        if tipo == "fecha":
            obtenido = a_fecha(obtenido)
        elif tipo == "numero":
            obtenido = a_numero(obtenido)
        else:
            obtenido = _plano(obtenido) if obtenido else None
            if obtenido == "no consta":
                obtenido = None

        if no_consta:
            desenlace = "abstención correcta" if obtenido is None else "INVENCIÓN"
        elif esperado is None:
            desenlace = "sin etiquetar"
        elif obtenido is None:
            desenlace = "omisión"
        elif obtenido == esperado:
            desenlace = "acierto"
        else:
            desenlace = "ERROR"
        out.append({"campo": col, "esperado": esperado, "obtenido": obtenido,
                    "desenlace": desenlace})
    return out


def medir(ruta, modo="determinista"):
    ref = leer_referencia(ruta)
    if not ref:
        print("No hay ninguna fila etiquetada (falta la columna «estado_correcto»).")
        return 1
    sistema, _ = leer_sistema([r["documento"] for r in ref], modo)

    ORDEN = ["acierto", "abstención correcta", "omisión", "ERROR", "INVENCIÓN",
             "sin etiquetar"]
    por_campo, por_desenlace, detalle = {}, dict.fromkeys(ORDEN, 0), []
    estados_ok = estados_no = 0

    for r in ref:
        s = sistema.get(r["documento"])
        for res in comparar(r, s):
            por_campo.setdefault(res["campo"], dict.fromkeys(ORDEN, 0))
            por_campo[res["campo"]][res["desenlace"]] += 1
            por_desenlace[res["desenlace"]] += 1
            if res["desenlace"] in ("ERROR", "INVENCIÓN", "omisión"):
                detalle.append((r["documento"], res))
        # El estado sólo se le exige a un documento sobre el que el evaluador no
        # se abstenga: si se abstuvo, ya lo ha dicho y no está afirmando nada.
        if s and not s.get("abstiene"):
            emitido = s["estado"]
            if emitido == r["estado"] or MAS_PRECISO_QUE.get(emitido) == r["estado"]:
                estados_ok += 1
            else:
                estados_no += 1
                detalle.append((r["documento"],
                                {"campo": "ESTADO", "esperado": r["estado"],
                                 "obtenido": emitido, "desenlace": "ERROR"}))

    ancho = max(len(c) for c in por_campo) + 2
    print(f"\nConjunto de referencia · {len(ref)} documento(s) etiquetado(s) · "
          f"modo {modo}\n")
    print(f"{'campo':{ancho}}" + "".join(f"{d[:11]:>13}" for d in ORDEN))
    print("-" * (ancho + 13 * len(ORDEN)))
    for campo, cuenta in por_campo.items():
        print(f"{campo:{ancho}}" + "".join(f"{cuenta[d]:>13}" for d in ORDEN))
    print("-" * (ancho + 13 * len(ORDEN)))
    print(f"{'TOTAL':{ancho}}" + "".join(f"{por_desenlace[d]:>13}" for d in ORDEN))

    afirmados = por_desenlace["acierto"] + por_desenlace["ERROR"]
    callados = por_desenlace["abstención correcta"] + por_desenlace["INVENCIÓN"]
    print(f"\n  Precisión de lo que afirma  "
          f"{100*por_desenlace['acierto']/afirmados:5.1f} %  "
          f"({por_desenlace['acierto']} de {afirmados} campos afirmados)"
          if afirmados else "\n  No afirma ningún campo.")
    if callados:
        print(f"  Prudencia cuando no consta  "
              f"{100*por_desenlace['abstención correcta']/callados:5.1f} %  "
              f"({por_desenlace['abstención correcta']} de {callados} campos que el "
              f"documento no dice)")
    if estados_ok + estados_no:
        print(f"  Estado de vigencia          "
              f"{100*estados_ok/(estados_ok+estados_no):5.1f} %  "
              f"({estados_ok} de {estados_ok+estados_no} documentos no abstenidos)")

    if detalle:
        print("\n  Dónde falla, uno por uno:")
        for doc, res in detalle:
            print(f"    · {doc[:38]:40} {res['campo']:20} {res['desenlace']:20} "
                  f"esperado {res['esperado']} · obtenido {res['obtenido']}")
    return 0


if __name__ == "__main__":
    args = [a for a in sys.argv[1:] if not a.startswith("--")]
    modo = "asistido" if "--asistido" in sys.argv else "determinista"
    sys.exit(medir(args[0] if args else "referencia/etiquetado.xlsx", modo))
